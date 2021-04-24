#!/usr/bin/python3
#
# written by @author ZyzonixDev
# published by ZyzonixDevelopments 
# -
# date      | 27/03/2021
# python-v  | 3.5.3
# -
# file      | monitoring.py
# project   | MontSy
# project-v | 0.9.6
# 
import sqlite3
import openpyxl
from datetime import datetime, date

def getDate():
    curDate = "" + str(date.today().strftime("%Y-%m-%d"))
    return curDate

def getTime():
    curTime = "" + str(datetime.now().strftime("%H:%M:%S"))
    return curTime

def sqlTableConstructor(functions):
    f_keys = functions.keys()
    command = "date timestamp, time timestamp"
    
    for key in f_keys:
        fout = functions[key]()
        generalkeys = list(fout.keys())
        keycount = len(generalkeys)
        counter = 1
        while counter <= keycount:
            res_count_key = list(fout[counter].keys())
            command += ", " + res_count_key[0] + " TEXT"
            counter += 1
    
    return command   

def xlsxConstructor(type, workbook, functions):
    f_keys = functions.keys()
    worksheet = workbook.create_sheet()
    worksheet.title = type
    REWmodulelist = []
    REWmodulelist.append("date")
    REWmodulelist.append("time")
    
    for key in f_keys:
        fout = functions[key]()
        generalkeys = list(fout.keys())
        keycount = len(generalkeys)
        counter = 1
        while counter <= keycount:
            res_count_key = list(fout[counter].keys())
            REWmodulelist.append(res_count_key[0])
            counter += 1
    
    for col, item in enumerate(REWmodulelist, start=1):
        worksheet.cell(1, col, item)

def createStorage(self, files, tables, modules, net_modules):
    from static import values
    filePathArray = []
    table_connection = {
        "generalmonitoring" : modules,
        "networkmonitoring" : net_modules       
    }
    
    for file in files:
        filePath = self.out_directory + "monitoring/" + file 
        
        if file.endswith(".db"):
            connection = sqlite3.connect(filePath)
            dbCursor = connection.cursor()
            #print(tables)
            for table in tables:
                command = sqlTableConstructor(table_connection[table])
                #print(command)
                dbCursor.execute("CREATE TABLE IF NOT EXISTS " + values.tables_names[table] + " (" + command + ")")
            connection.commit()
            connection.close()
        
        else:
            wbFile = openpyxl.Workbook()
            for table in tables:
                xlsxConstructor(values.tables_names[table], wbFile, table_connection[table])
            std_sheet = wbFile.get_sheet_by_name("Sheet")
            wbFile.remove_sheet(std_sheet)
            wbFile.save(filename=filePath)
            wbFile.close()
        filePathArray.append(filePath)
    
    # returning file paths for further editing
    return filePathArray

# retrieving modules; adding to dict
def retrieveModules(self):
    modules = {}
    net_modules = {}
    fromConfig = self.enabled_module_list["m"]
    # getting all available functions 
    # access functions: modules[function]()
    keys = fromConfig.keys()
    import modules.mon as mon
    
    for key in keys:
        if fromConfig[key].startswith("net_"):
            net_modules[key] = getattr(mon, fromConfig[key])
        else:    
            modules[key] = getattr(mon, fromConfig[key])
    
    return modules, net_modules
