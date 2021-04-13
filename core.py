#!/usr/bin/python3
#
# written by @author ZyzonixDev
# published by ZyzonixDevelopments 
# -
# date      | 18/03/2021
# python-v  | 3.5.3
# -
# file      | core.py
# project   | MontSy
# project-v | 0.9.1 (beta)
# 
from datetime import date, datetime
from configparser import ConfigParser
import sqlite3
import traceback
import os
import sys
import threading
import openpyxl
import static.values as values
from static.values import colorsetting as color
from modules import overview, monitoring


# writing console output to console and logfile
class LogWriter(object):
    def __init__(self, *files):
        # retrieving files / output locations
        self.files = files

    def write(self, obj):
        # getting files (logfile / console (as file))
        for file in self.files:
            file.write(obj)
            file.flush()

    def flush(self):
        # flushing written lines/files
        for file in self.files:
            file.flush()


# software functionalities (sov and mon)
class Handlers(object):
    
    # system overview handler | key: s
    def handleOverview(self):
        output = overview.prepareStoring(self)
        module_list = overview.retrieveModules(self)

        try:
            for key in module_list:
                data = module_list[key]()

                # checking if function returns nothing
                if not data:
                    print(self.getTime(), "empty - " + key)
                else:
                    print(self.getTime(), "writing data - " + key)
                    output.write(values.sov_headers[key])

                    for dictio in data:
                        d_keys = dictio.keys()

                        for d_key in d_keys:
                            # kicking values if they're empty 
                            if dictio[d_key]:
                                output.write("\n" + d_key + " " + str(dictio[d_key]))
            
            print(self.getTime(), color.GREEN + "created system overview successfully\n" + color.END)            
            output.close()
        
        except:
            traceback.print_exc()
            print(self.getTime(), color.RED + "something went wrong - wasn't able to create a system overview" + color.END)


    # system monitoring handler | key: m 
    def handleMonitoring(self, files, modules, net_modules): 
        # modules = dict[number : modulename]
        print(self.getTime(), "collecting data and writing to files")
        # REWORK: SPLIT NETWORK AND NORMAL MODULES DURING RETRIEVE
        connectionArray = {}
        xlsxfile = ""
        
        for file in files:
            if file.endswith(".db"):
                connection = sqlite3.connect(file)
                connectionArray["sqlite"] = connection
            elif file.endswith(".xlsx"):
                workbook = openpyxl.load_workbook(file)
                connectionArray["xlsx"] = workbook
                xlsxfile = file
        try:
            data = {}
            net_data = {}
            cmd = ""
            net_cmd = ""
            
            if modules:
                cmd += "'" + str(monitoring.getDate()) + "', '" + str(monitoring.getTime()) + "'"
                counter = 1
                for module in modules:
                    sdata = modules[module]()
                    for key in sdata:
                        selected_data = sdata[key]
                        for smod in selected_data:
                            data[counter] = selected_data[smod]
                            cmd += ", " + str(selected_data[smod])
                            counter += 1
            
            if net_modules:
                net_cmd += "'" + str(monitoring.getDate()) + "', '" + str(monitoring.getTime()) + "'"
                counter = 1
                for net_module in net_modules:
                    sdata = net_modules[net_module]()
                    for key in sdata:
                        selected_data = sdata[key]
                        for smod in selected_data:
                            net_data[counter] = selected_data[smod]
                            net_cmd += ", '" + str(selected_data[smod]) + "'"
                            counter += 1
            
            for connectionKey in connectionArray.keys():
                if connectionKey == "sqlite":
                    connection = connectionArray[connectionKey]
                    dbCursor = connection.cursor()
                    if cmd != "":
                        #print(cmd)
                        dbCursor.execute("INSERT INTO General VALUES (" + cmd + ")")
                    if net_cmd != "":
                        #print(net_cmd)
                        dbCursor.execute("INSERT INTO Network VALUES (" + net_cmd + ")")
                    connection.commit()
                    connection.close()
                
                elif connectionKey == "xlsx":
                    workbook = connectionArray[connectionKey]
                    
                    if cmd != "":
                        worksheet = workbook.get_sheet_by_name("General")
                        row = workbook.active.max_row + 1
                        data_list = []
                        data_list.append(monitoring.getDate())
                        data_list.append(monitoring.getTime())
                        for key in data.keys():
                            data_list.append(data[key])
                        for column, item in enumerate(data_list, start=1):
                            worksheet.cell(row, column, item)
                    
                    if net_cmd != "":
                        worksheet = workbook.get_sheet_by_name("Network")
                        row = workbook.active.max_row + 1
                        data_list = []
                        data_list.append(monitoring.getDate())
                        data_list.append(monitoring.getTime())
                        for key in net_data.keys():
                            data_list.append(net_data[key])
                        for column, item in enumerate(data_list, start=1):
                            worksheet.cell(row, column, item)
                    workbook.save(filename=xlsxfile)
                    workbook.close()
            
            # scheduling next run
            threading.Timer(self.MES_TIME, Handlers.handleMonitoring, [self, files, modules, net_modules]).start()
        
        except KeyboardInterrupt as e:
            print(e)
            print(self.getTime(), color.RED + "stopped Montsy"  + color.END)
                        

    # system monitoring handler (preparation) | key: m 
    def prepareMonitoring(self):
        module_list = self.enabled_module_list["m"]
        # general: everything except network
        sheetNeeded = {
            "generalmonitoring" : False,
            "networkmonitoring" : False
        }
        keys = sheetNeeded.keys()
        mod_keys = module_list.keys()
        
        for key in mod_keys:
            selected_module = module_list[key]
            if selected_module.startswith("net_"):
                sheetNeeded["networkmonitoring"] = True
            else:
                sheetNeeded["generalmonitoring"] = True
        
        fileName = str(date.today()) + "_" + str(datetime.now().strftime("%H-%M-%S"))        
        files = []
        fileTypes = values.storageMethod[self.storage_m]
        
        for type in fileTypes:
            files.append(fileName + type)
        
        # returns created files (not path/name)
        prepTableNames = list(keys)
        for key in prepTableNames:
            if not sheetNeeded[key]:
                prepTableNames.remove(key)    
        
        # collecting and summing up required data
        modules, net_modules = monitoring.retrieveModules(self)
        fileArray = monitoring.createStorage(self, files, prepTableNames, modules, net_modules)
        if not fileArray: return
        Handlers.handleMonitoring(self, fileArray, modules, net_modules)


# core class (__init__ initializes the system)
class Core(object):
    # console time service
    def getTime(self):
        curTime = "[" + str(datetime.now().strftime("%H:%M:%S")) + "]"
        return curTime

    # getting configuration from config file
    def getModConfig(self, configImport, importModules):
        modulespec = self.static[importModules]
        modulelist = {}
        try:
            print(self.getTime(), "getting configuration for '" + modulespec[importModules] + "'")
            modulesToImport = configImport[modulespec["config"]]
            
            if modulesToImport:
                counter = 1
                
                for module in modulesToImport:
                    # selecting all modules where the value == True (boolean not string)
                    selected_module = configImport.getboolean(modulespec["config"], module)
                    if selected_module == True:
                        modulelist[counter] = module
                        print(self.getTime(), "moduleinfo (" + modulespec[importModules] + ") - enabled - installed:", module)
                    else:
                        print(self.getTime(), "moduleinfo (" + modulespec[importModules] + ") - disabled:", module)
                    counter += 1
            
            else:
                print(self.getTime(),"no modules found - check config file")
                return False
            # saving collected modules to list    
            self.enabled_module_list[importModules] = modulelist 
            if not modulelist: return False
            print(self.getTime(), color.GREEN + "reading " + modulespec[importModules] + "-configuration passed\n" + color.END)
            return True
        
        except:
            print(self.getTime(), color.RED + "something went wrong - was not able to get the configuration - exiting...\n" + color.END)
            traceback.print_exc()
            return False

    # writing console to log
    def writeLog(self):
        self.logFile = open(os.getcwd() + "/logs/" + str(date.today()) + "_" + str(
            datetime.now().strftime("%H-%M-%S")) + "_log.txt", "w")
        sys.stdout
        sys.stdout = LogWriter(sys.stdout, self.logFile)    

    def __init__(self):
        print("\n" + self.getTime(), "running " + color.RED + color.BOLD +  "Mon(t)Sy" + color.END + " (system monitoring and overview) application \n")
        # saving ressources global (like public) // retrieving missing from config file
        configImport = ConfigParser()
        configFile = os.getcwd() + "/static/config.ini"
        # os.getcwd() returns execution directory
        configImport.read(configFile)
        # preparing instance for saving execution directory
        configWriter = ConfigParser(comment_prefixes='/', allow_no_value=True)
        configWriter.read(configFile)
        
        try:
            # importing config data
            self.baseFilePath = os.getcwd() + "/"
            configWriter["CONFIGURATION"]["basefilepath"] = self.baseFilePath
            # writing execution directory to file
            with open(configFile, "w") as file:
                configWriter.write(file)
            self.MES_TIME = int(configImport["CONFIGURATION"]["log_dur"])
            self.storage_m = int(configImport["CONFIGURATION"]["mon_storage_method"])
            self.out_directory = configImport["CONFIGURATION"]["output_dest"]
        except:
            traceback.print_exc()
            return

        # initializing log, if enabled
        if int(configImport["CONFIGURATION"]["log_enabled"]) == 1:
            self.writeLog()

        # preparing lists for enabled modules
        self.enabled_module_list = values.module_dictionary
        # importing preset settings
        self.static = values.static
        # starting software
        print(self.getTime(), "reading arguments (" + str(len(sys.argv)) + ")")
        
        # if a valid argument is given --> selected method will be executed
        if len(sys.argv) != 1:
            prov_arg = sys.argv[1]
            if not prov_arg in self.static.keys():
                print(self.getTime(), "wrong argument provided (" + prov_arg + ") - exiting...\n")
                return
            print(self.getTime(), self.static[prov_arg]["console"])

            # importing modules (checking config for en- and disabled modules)
            if self.getModConfig(configImport, prov_arg) == False:
                print(self.getTime(), color.RED + "all modules are disabled - exiting... " + color.END)
                return
            
            # starting selected module
            if prov_arg == "s":  
                Handlers.handleOverview(self),
            elif prov_arg == "m":
                Handlers.prepareMonitoring(self)

        # if no argument provided: systemooverview will run first, then initializing the monitoring part
        else:
            # getting configuration for both possibilities
            getModules = "all"
            getAll = ["s","m"]
            check = []
            print(self.getTime(), self.static[getModules][getModules])
            
            for method in getAll:
                if self.getModConfig(configImport, method) == False:
                    print(self.getTime(), color.RED + "all modules for method '" + self.static[method][method] + "' are disabled" + color.END)
                    check.append("PLACEHOLDER")
                if len(check) == 2:
                    print(self.getTime(), color.RED + "all modules are disabled - please change the config file (static/config.ini) - exiting..." + color.END)
                    return    
                # prints module list
                #print(self.enabled_module_list[method])
            
            print(self.getTime(), "starting to create a system overview")
            # func for sov-handling here
            Handlers.handleOverview(self)
            print(self.getTime(), "installing system monitoring")
            # func for mon-handling here
            Handlers.prepareMonitoring(self)

# initializes the core system
if __name__ == "__main__":
    Core()